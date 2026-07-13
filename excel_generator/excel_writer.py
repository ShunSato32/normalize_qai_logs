import os
import json
import openpyxl
from openpyxl.styles import Alignment, Border, Side, PatternFill, Font
from openpyxl.utils import get_column_letter
from datetime import datetime
from typing import List, Dict, Any

CIRCLED_NUMS = ["①", "②", "➂", "④", "⑤", "⑥", "⑦", "⑧", "➈", "➉"]

def get_circled_num_doc_content(rank: int) -> str:
    if rank == 3:
        return "➂"  # U+2782
    elif rank == 9:
        return "➈"  # U+2788
    elif rank == 10:
        return "➉"  # U+2789
    else:
        return CIRCLED_NUMS[rank - 1]

def get_circled_num_score(rank: int) -> str:
    if rank == 3:
        return "③"  # U+2462
    elif rank == 9:
        return "➈"  # U+2788
    elif rank == 10:
        return "➉"  # U+2789
    else:
        return CIRCLED_NUMS[rank - 1]

# Build the complete template headers list representing the original 110 columns
TEMPLATE_HEADERS = [
    # 15.1
    "output_row_id",
    "No.",
    "interaction_id",
    "message_id",
    "conversation_id",
    "reset_session_id",
    "reset_session_no",
    "interaction_no_in_session",
    "is_first_interaction_row",
    # 15.2
    "user_name",
    "team_name",
    "質問内容",
    "user_content_raw",
    "is_natural_language_query",
    "is_reset_request",
    "is_category_selection",
    "is_system_command",
    # 15.3
    "started_at_utc",
    "completed_at_utc",
    "started_at_jst",
    "completed_at_jst",
    "latency_sec",
    # 15.4
    "selected_function",
    "predicted_category",
    "user_selected_category",
    "fist_category",
    "final_category",
    "category_source",
    "is_unclassified",
    # 15.5
    "回答内容",
    "has_answer",
    "is_no_answer",
    "assistant_event_count",
    "has_error",
    "error_count",
    "error_message",
    # 15.6
    "retrieval_count",
    "retrieval_stored_count",
    "retrieval_truncated",
]

# Retrieval fixed columns (Columns 41 to 70)
for k in range(1, 11):
    doc_c = get_circled_num_doc_content(k)
    score_c = get_circled_num_score(k)
    TEMPLATE_HEADERS.append(f"参照文書{doc_c}")
    TEMPLATE_HEADERS.append(f"参考箇所{doc_c}")
    TEMPLATE_HEADERS.append(f"score{score_c}")

TEMPLATE_HEADERS.extend([
    # 15.8
    "feedback_id",
    "feedback_seq",
    "feedback_count",
    "feedback_rating",
    "feedback_comment",
    "feedback_at_utc",
    "feedback_at_jst",
    "is_latest_feedback",
    # 15.9
    "session_ends_with_reset",
    "source_files",
    "source_event_row_count",
    "source_event_types",
    "normalization_version",
    "normalization_warnings",
    # Existing Business / Evaluation columns
    "質問/回答\n分類",
    "対象/\n対象外",
])

for k in range(1, 11):
    doc_c = get_circled_num_doc_content(k)
    TEMPLATE_HEADERS.append(f"参照箇所{doc_c}")

TEMPLATE_HEADERS.extend([
    "教師データ有無",
    "キーワード1",
    "キーワード2",
    "キーワード3",
    "Hit判定",
    "回答判定",
    "正答判定",
    "分析コメント",
    "原因",
    "対応方針",
    "対応進捗",
    "Vantiq相談",
    "日付",
    "所属",
    "messageID"
])

# Renames mapping from TEMPLATE_HEADERS to English equivalent physical_names
RENAMED_COLUMNS = {
    "output_row_id": "unique_row_id",
    "質問内容": "question",
    "回答内容": "answer",
    "質問/回答\n分類": "qa_classification",
    "対象/\n対象外": "is_target",
    "教師データ有無": "has_training_data",
    "キーワード1": "keyword_1",
    "キーワード2": "keyword_2",
    "キーワード3": "keyword_3",
    "Hit判定": "hit_judgment",
    "回答判定": "answer_judgment",
    "正答判定": "accuracy_judgment",
    "分析コメント": "analysis_comment",
    "原因": "error_cause",
    "対応方針": "action_plan",
    "対応進捗": "progress",
    "Vantiq相談": "vantiq_consultation",
    "日付": "date_jst",
    "所属": "department",
}
for k in range(1, 11):
    doc_c = get_circled_num_doc_content(k)
    score_c = get_circled_num_score(k)
    RENAMED_COLUMNS[f"参照文書{doc_c}"] = f"ref_doc_{k}"
    RENAMED_COLUMNS[f"参考箇所{doc_c}"] = f"ref_text_{k}"
    RENAMED_COLUMNS[f"score{score_c}"] = f"score_{k}"
    RENAMED_COLUMNS[f"参照箇所{doc_c}"] = f"ref_check_{k}"

# Build HEADERS by copying TEMPLATE_HEADERS but removing UTC columns and applying renames
DELETED_COLUMNS = {"started_at_utc", "completed_at_utc"}
HEADERS = []
for h in TEMPLATE_HEADERS:
    if h not in DELETED_COLUMNS:
        HEADERS.append(RENAMED_COLUMNS.get(h, h))

WRAP_COLS = {
    "question", "user_content_raw", "answer", "error_message", 
    "feedback_comment", "normalization_warnings", "analysis_comment", "evaluation_comment"
}
for k in range(1, 11):
    WRAP_COLS.add(f"ref_text_{k}")

def parse_iso_datetime(dt_str: str) -> datetime:
    if not dt_str:
        return None
    try:
        # standard ISO string parsing, strip offset (e.g. +09:00, +00:00) to keep naive datetime local to that tz
        if '+' in dt_str:
            dt_str = dt_str.split('+')[0]
        if dt_str.endswith('Z'):
            dt_str = dt_str[:-1]
        return datetime.fromisoformat(dt_str)
    except Exception:
        return None

def load_target_categories(config_dir: str = "config") -> List[Dict[str, Any]]:
    path = os.path.join(config_dir, "target_categories.json")
    if not os.path.exists(path):
        # Try finding in parent directory (project root)
        project_root = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
        alt_path = os.path.join(project_root, config_dir, "target_categories.json")
        if os.path.exists(alt_path):
            path = alt_path
        else:
            return []
    try:
        with open(path, "r", encoding="utf-8") as f:
            return json.load(f)
    except Exception:
        return []

def flatten_category_tree(categories_json: List[Dict[str, Any]]):
    top_level_names = []
    all_tree_names = []
    node_map = {}

    def traverse(node, top_level_name, current_path):
        name = node.get("name", "")
        if not name:
            return
        path_str = " > ".join(current_path + [name]) if current_path else name
        all_tree_names.append(path_str)
        info = {"name": name, "full_path": path_str, "top_level": top_level_name}
        node_map[name] = info
        node_map[path_str] = info
        node_map["/".join(current_path + [name])] = info
        node_map[" | ".join(current_path + [name])] = info
        for child in node.get("children", []):
            traverse(child, top_level_name, current_path + [name])

    for root_node in categories_json:
        r_name = root_node.get("name", "")
        if r_name:
            top_level_names.append(r_name)
            traverse(root_node, r_name, [])

    return top_level_names, all_tree_names, node_map

def aggregate_daily_summary(rows: List[Dict[str, Any]], top_level_names: List[str], all_tree_names: List[str], node_map: Dict[str, Dict[str, str]]):
    daily_rows: Dict[str, List[Dict[str, Any]]] = {}
    for r in rows:
        dt_str = r.get("started_at_jst", "")
        dt = parse_iso_datetime(dt_str)
        if dt:
            day_str = dt.strftime("%Y/%m/%d")
        else:
            if dt_str and len(dt_str) >= 10:
                day_str = dt_str[:10].replace("-", "/")
            else:
                day_str = "日付不明"
        if day_str not in daily_rows:
            daily_rows[day_str] = []
        daily_rows[day_str].append(r)

    sorted_days = sorted(daily_rows.keys())
    results = []
    for day_str in sorted_days:
        day_r = daily_rows[day_str]
        q_rows = [r for r in day_r if (str(r.get("is_first_interaction_row")) in ("1", "True", "true") or r.get("is_first_interaction_row") == 1) and r.get("qa_classification") != "④集計対象外"]
        if not q_rows and day_r:
            q_rows = [r for r in day_r if r.get("qa_classification") != "④集計対象外"]
        q_count = len(q_rows)
        
        users = set(r.get("user_name", "") for r in q_rows if r.get("user_name"))
        user_count = len(users)
        
        c1 = len([r for r in q_rows if r.get("qa_classification") == "①該当無"])
        c5 = len([r for r in q_rows if r.get("qa_classification") == "⑤その他"])
        
        bad_cnt = 0
        good_cnt = 0
        unset_cnt = 0
        for r in q_rows:
            fb = str(r.get("feedback_rating", "")).strip().lower()
            if fb == "bad":
                bad_cnt += 1
            elif fb == "good":
                good_cnt += 1
            else:
                unset_cnt += 1

        top_counts = {name: 0 for name in top_level_names}
        tree_counts = {name: 0 for name in all_tree_names}

        for r in q_rows:
            cat = (r.get("final_category") or r.get("user_selected_category") or 
                   r.get("predicted_category") or r.get("fist_category") or "")
            if not cat:
                if "未分類" in node_map:
                    cat = "未分類"
                else:
                    continue
            info = node_map.get(cat)
            if not info:
                for sep in [" > ", " >", "> ", ">", "/", " | ", "｜"]:
                    if sep in cat:
                        leaf = cat.split(sep)[-1].strip()
                        if leaf in node_map:
                            info = node_map[leaf]
                            break
            if not info and cat in node_map:
                info = node_map[cat]
            if info:
                t_level = info.get("top_level", "")
                f_path = info.get("full_path", "")
                if t_level in top_counts:
                    top_counts[t_level] += 1
                if f_path in tree_counts:
                    tree_counts[f_path] += 1

        results.append({
            "day": day_str,
            "q_count": q_count,
            "user_count": user_count,
            "c1": c1,
            "c5": c5,
            "bad": bad_cnt,
            "good": good_cnt,
            "unset": unset_cnt,
            "top_counts": top_counts,
            "tree_counts": tree_counts
        })

    return results

def populate_summary_sheets(wb: openpyxl.Workbook, rows: List[Dict[str, Any]]):
    cats_json = load_target_categories()
    top_names, tree_names, node_map = flatten_category_tree(cats_json)
    added_headers = top_names + tree_names
    daily_results = aggregate_daily_summary(rows, top_names, tree_names, node_map)

    # 1. 集計詳細 (O列 / col 15 以降)
    if "集計詳細" in wb.sheetnames:
        ws_dt = wb["集計詳細"]
        ref_cell = ws_dt.cell(row=2, column=14)
        for idx, h_name in enumerate(added_headers):
            col_idx = 15 + idx
            cell = ws_dt.cell(row=2, column=col_idx, value=h_name)
            if ref_cell and ref_cell.font:
                cell.font = Font(name=ref_cell.font.name, size=ref_cell.font.size, bold=ref_cell.font.bold, color=ref_cell.font.color)
            if idx < len(top_names):
                cell.fill = PatternFill(fill_type="solid", start_color="A9D08E", end_color="A9D08E")
            else:
                cell.fill = PatternFill(fill_type="solid", start_color="9BC2E6", end_color="9BC2E6")
            if ref_cell and ref_cell.border:
                cell.border = Border(left=ref_cell.border.left, right=ref_cell.border.right, top=ref_cell.border.top, bottom=ref_cell.border.bottom)
            if ref_cell and ref_cell.alignment:
                cell.alignment = Alignment(horizontal="center", vertical=ref_cell.alignment.vertical, wrap_text=True)

        for r_idx, d_res in enumerate(daily_results, start=3):
            ws_dt.cell(row=r_idx, column=1, value=d_res["day"])
            ws_dt.cell(row=r_idx, column=2, value=d_res["q_count"])
            ws_dt.cell(row=r_idx, column=3, value=d_res["user_count"])
            ws_dt.cell(row=r_idx, column=4, value=d_res["c1"])
            ws_dt.cell(row=r_idx, column=5, value=d_res["c5"])
            ws_dt.cell(row=r_idx, column=6, value=f"=B{r_idx}")
            ws_dt.cell(row=r_idx, column=7, value=f"=IF(F{r_idx}=0,0,(F{r_idx}-D{r_idx})/F{r_idx})")
            ws_dt.cell(row=r_idx, column=8, value=d_res["bad"])
            ws_dt.cell(row=r_idx, column=9, value=d_res["good"])
            ws_dt.cell(row=r_idx, column=10, value=d_res["unset"])
            ws_dt.cell(row=r_idx, column=11, value=f"=SUM(H{r_idx}:J{r_idx})")
            ws_dt.cell(row=r_idx, column=12, value=f"=SUM(H{r_idx}:I{r_idx})")
            ws_dt.cell(row=r_idx, column=13, value=f"=IF(B{r_idx}=0,0,L{r_idx}/B{r_idx})")
            ws_dt.cell(row=r_idx, column=14, value=f"=IF(L{r_idx}=0,0,I{r_idx}/L{r_idx})")

            for h_idx, top_name in enumerate(top_names):
                col_idx = 15 + h_idx
                ws_dt.cell(row=r_idx, column=col_idx, value=d_res["top_counts"].get(top_name, 0))
            for h_idx, tree_name in enumerate(tree_names):
                col_idx = 15 + len(top_names) + h_idx
                ws_dt.cell(row=r_idx, column=col_idx, value=d_res["tree_counts"].get(tree_name, 0))

            for col_i in (7, 13, 14):
                c = ws_dt.cell(row=r_idx, column=col_i)
                c.number_format = '0.0%'

    # 2. 集計概要 (K列 / col 11 以降は第１カテゴリのグルーピング集計値だけ)
    if "集計概要" in wb.sheetnames:
        ws_ov = wb["集計概要"]
        ref_cell = ws_ov.cell(row=2, column=10)
        for idx, h_name in enumerate(top_names):
            col_idx = 11 + idx
            cell = ws_ov.cell(row=2, column=col_idx, value=h_name)
            if ref_cell and ref_cell.font:
                cell.font = Font(name=ref_cell.font.name, size=ref_cell.font.size, bold=ref_cell.font.bold, color=ref_cell.font.color)
            cell.fill = PatternFill(fill_type="solid", start_color="A9D08E", end_color="A9D08E")
            if ref_cell and ref_cell.border:
                cell.border = Border(left=ref_cell.border.left, right=ref_cell.border.right, top=ref_cell.border.top, bottom=ref_cell.border.bottom)
            if ref_cell and ref_cell.alignment:
                cell.alignment = Alignment(horizontal="center", vertical=ref_cell.alignment.vertical, wrap_text=True)

        for r_idx, d_res in enumerate(daily_results, start=3):
            ws_ov.cell(row=r_idx, column=1, value=d_res["day"])
            ws_ov.cell(row=r_idx, column=2, value=f"=集計詳細!B{r_idx}")
            ws_ov.cell(row=r_idx, column=3, value=f"=集計詳細!C{r_idx}")
            ws_ov.cell(row=r_idx, column=4, value=f"=集計詳細!G{r_idx}")
            ws_ov.cell(row=r_idx, column=5, value=f"=集計詳細!H{r_idx}")
            ws_ov.cell(row=r_idx, column=6, value=f"=集計詳細!I{r_idx}")
            ws_ov.cell(row=r_idx, column=7, value=f"=集計詳細!J{r_idx}")
            ws_ov.cell(row=r_idx, column=8, value=f"=E{r_idx}+F{r_idx}")
            ws_ov.cell(row=r_idx, column=9, value=f"=IF(B{r_idx}=0,0,H{r_idx}/B{r_idx})")
            ws_ov.cell(row=r_idx, column=10, value=f"=IF(H{r_idx}=0,0,F{r_idx}/H{r_idx})")

            for h_idx, top_name in enumerate(top_names):
                col_idx = 11 + h_idx
                ws_ov.cell(row=r_idx, column=col_idx, value=f"=集計詳細!{get_column_letter(15 + h_idx)}{r_idx}")

            for col_i in (4, 9, 10):
                c = ws_ov.cell(row=r_idx, column=col_i)
                c.number_format = '0.0%'

def write_integrated_to_excel(template_path: str, output_path: str, rows: List[Dict[str, Any]]):
    # Load template workbook
    try:
        wb = openpyxl.load_workbook(template_path, data_only=False)
    except Exception as e:
        if "BadZipFile" in type(e).__name__ or "not a zip file" in str(e).lower():
            print(f"\n[エラー: テンプレートExcelファイルが破損しているか、Git LFSのポインターファイルのままです]", file=sys.stderr)
            print(f"対象ファイル: {template_path}", file=sys.stderr)
            print(f"【原因と対処手順】", file=sys.stderr)
            print(f"  1. Git LFS (Large File Storage) ポインターの可能性: ", file=sys.stderr)
            print(f"     別PCで 'git lfs pull' または 'git lfs install' を実行して実体ファイル(約993KB)を取得するか、", file=sys.stderr)
            print(f"     元のPCから実体の【社名】実施記録分析シート.xlsxをコピーして上書き配置してください。", file=sys.stderr)
            print(f"  2. Windowsの改行コード自動変換 (autocrlf) / IRM暗号化の可能性: ", file=sys.stderr)
            print(f"     Git clone時にバイナリ破損しないよう .gitattributes が導入されました。ファイルが暗号化されていないかもご確認ください。\n", file=sys.stderr)
        raise
    
    if "実施記録シート" not in wb.sheetnames:
        raise ValueError("Template Excel file must contain a worksheet named '実施記録シート'")
        
    ws = wb["実施記録シート"]
    
    # 1. Cache the styles from row 4 before clearing values.
    # We map from physical column name to its style, based on original TEMPLATE_HEADERS mapping.
    physical_styles_cache = {}
    for idx, orig_header in enumerate(TEMPLATE_HEADERS):
        col_idx = idx + 2
        if col_idx <= ws.max_column:
            cell = ws.cell(row=4, column=col_idx)
            style_dict = {
                "font": cell.font,
                "fill": cell.fill,
                "border": cell.border,
                "alignment": cell.alignment,
                "number_format": cell.number_format
            }
            physical_styles_cache[orig_header] = style_dict
            new_name = RENAMED_COLUMNS.get(orig_header)
            if new_name:
                physical_styles_cache[new_name] = style_dict

    # 2. Load column configuration dynamically
    script_dir = os.path.dirname(os.path.abspath(__file__))
    project_root = os.path.dirname(script_dir)
    config_paths_to_try = [
        os.path.join(project_root, "config", "column_config.json"),
        os.path.join(script_dir, "config", "column_config.json"),
        "config/column_config.json"
    ]
    config_path = None
    for p in config_paths_to_try:
        if os.path.exists(p):
            config_path = p
            break

    active_columns = []
    if config_path and os.path.exists(config_path):
        try:
            with open(config_path, "r", encoding="utf-8") as f:
                config_data = json.load(f)
            for item in config_data:
                pname = item.get("physical_name")
                jname = item.get("japanese_name") or pname
                if pname:
                    active_columns.append((pname, jname))
        except Exception as e:
            import sys
            print(f"Warning: Failed to load column_config.json: {e}. Using default headers.", file=sys.stderr)
            
    if not active_columns:
        active_columns = [(h, h) for h in HEADERS]
        
    # Map physical name to its active column letter in the JST sheet
    col_letters = {}
    for idx, (pname, jname) in enumerate(active_columns):
        col_idx = idx + 2
        col_letters[pname] = get_column_letter(col_idx)
        
    # 3. Clear all values from row 4 to the end of the sheet
    max_row = ws.max_row
    max_col = ws.max_column
    for r in range(4, max_row + 1):
        for c in range(1, max_col + 1):
            ws.cell(row=r, column=c).value = None
            
    # 4. Overwrite Row 3 headers from Column B (index 2) onwards
    for idx, (pname, jname) in enumerate(active_columns):
        col_idx = idx + 2
        cell = ws.cell(row=3, column=col_idx)
        cell.value = jname
    # Clear any leftover headers beyond the active columns
    for col_idx in range(2 + len(active_columns), max_col + 1):
        ws.cell(row=3, column=col_idx).value = None
        
    # 5. Populate rows from row 4 onwards
    last_row_index = 3 + len(rows)
    if len(rows) == 0:
        last_row_index = 4
        
    for row_num, row_data in enumerate(rows, start=4):
        is_sys_cmd = (row_data.get("is_system_command") in (1, "1", True, "true", "True"))
        for idx, (pname, jname) in enumerate(active_columns):
            col_idx = idx + 2
            cell = ws.cell(row=row_num, column=col_idx)
            
            # Write value or formula
            if pname == "No.":
                cell.value = f"=ROW(A{row_num})-3"
            elif pname == "qa_classification":
                val = row_data.get("qa_classification", "")
                if is_sys_cmd and not val:
                    cell.value = "④集計対象外"
                else:
                    cell.value = val
            elif pname == "is_target":
                if is_sys_cmd:
                    cell.value = "×"
                else:
                    # References "qa_classification" (CC in default)
                    q_a_class_col = col_letters.get("qa_classification", "CC")
                    cell.value = f'=IF(OR(LEFT({q_a_class_col}{row_num},1)="①", LEFT({q_a_class_col}{row_num},1)="⑤"), "⚪︎", "×")'
            elif pname.startswith("ref_check_") and len(pname) > 10:
                try:
                    rank = int(pname[10:])
                except ValueError:
                    rank = 1
                ref_desc_col = col_letters.get(f"ref_text_{rank}", "AP")
                kw1_col = col_letters.get("keyword_1", "CT")
                kw2_col = col_letters.get("keyword_2", "CU")
                kw3_col = col_letters.get("keyword_3", "CV")
                cell.value = f'=IF(OR(AND(${kw1_col}{row_num}<>"",ISNUMBER(SEARCH(${kw1_col}{row_num}, {ref_desc_col}{row_num}))),AND(${kw2_col}{row_num}<>"",ISNUMBER(SEARCH(${kw2_col}{row_num}, {ref_desc_col}{row_num}))),AND(${kw3_col}{row_num}<>"",ISNUMBER(SEARCH(${kw3_col}{row_num}, {ref_desc_col}{row_num})))),"〇","-")'
            elif pname == "hit_judgment":
                check_cols = []
                for k in range(1, 11):
                    col_let = col_letters.get(f"ref_check_{k}")
                    if col_let:
                        check_cols.append(f'{col_let}{row_num}="〇"')
                if check_cols:
                    cell.value = f'=IF(OR({",".join(check_cols)}),"〇","")'
                else:
                    cell.value = ""
            elif pname == "date_jst":
                start_jst_col = col_letters.get("started_at_jst", "U")
                cell.value = f'=TEXT({start_jst_col}{row_num},"yyyy/mm/dd")'
            elif pname == "department":
                user_name_col = col_letters.get("user_name", "K")
                cell.value = f'=VLOOKUP({user_name_col}{row_num},社員マスタ!C:D,2,0)'
            elif pname == "messageID":
                cell.value = str(row_data.get("message_id", ""))
                cell.data_type = "s"
            elif pname in ("unique_row_id", "interaction_id", "message_id", "conversation_id", "reset_session_id", "feedback_id"):
                cell.value = str(row_data.get(pname, ""))
                cell.data_type = "s"
            elif pname in ("started_at_utc", "completed_at_utc", "started_at_jst", "completed_at_jst", "feedback_at_utc", "feedback_at_jst"):
                val_str = row_data.get(pname, "")
                parsed_dt = parse_iso_datetime(val_str)
                cell.value = parsed_dt
                cell.number_format = 'yyyy/mm/dd hh:mm:ss'
            elif pname == "latency_sec":
                val = row_data.get(pname, "")
                if val != "" and val is not None:
                    cell.value = float(val)
                else:
                    cell.value = None
            elif pname.startswith("score_") and len(pname) > 6:
                try:
                    rank = int(pname[6:])
                except Exception:
                    rank = 1
                col_name = f"retrieval_{rank:02d}"
                json_str = row_data.get(col_name)
                if json_str:
                    try:
                        data = json.loads(json_str)
                        score = data.get("score")
                        if score is not None and score != "":
                            cell.value = float(score)
                            cell.number_format = '0.000000'
                        else:
                            cell.value = None
                    except Exception:
                        cell.value = None
                else:
                    cell.value = None
            elif pname.startswith("ref_doc_") and len(pname) > 8:
                try:
                    rank = int(pname[8:])
                except Exception:
                    rank = 1
                col_name = f"retrieval_{rank:02d}"
                json_str = row_data.get(col_name)
                if json_str:
                    try:
                        data = json.loads(json_str)
                        filename = data.get("filename") or data.get("fileName") or data.get("display_name") or ""
                        cell.value = filename
                    except Exception:
                        cell.value = ""
                else:
                    cell.value = ""
            elif pname.startswith("ref_text_") and len(pname) > 9:
                try:
                    rank = int(pname[9:])
                except Exception:
                    rank = 1
                col_name = f"retrieval_{rank:02d}"
                json_str = row_data.get(col_name)
                if json_str:
                    try:
                        data = json.loads(json_str)
                        cell.value = data.get("content", "")
                    except Exception:
                        cell.value = ""
                else:
                    cell.value = ""
            elif pname in (
                "feedback_seq", "feedback_count", "retrieval_count", "retrieval_stored_count", 
                "retrieval_truncated", "is_first_interaction_row", "is_latest_feedback", 
                "is_natural_language_query", "is_reset_request", "is_category_selection", 
                "is_system_command", "session_ends_with_reset", "has_answer", "is_no_answer", 
                "assistant_event_count", "has_error", "error_count", "source_event_row_count"
            ):
                val = row_data.get(pname)
                if val != "" and val is not None:
                    cell.value = int(val)
                else:
                    cell.value = 0
            else:
                key_map = {
                    "predicted_category": "predicted_category",
                    "user_selected_category": "user_selected_category",
                    "answer": "answer",
                    "fist_category": "fist_category",
                    "final_category": "final_category",
                    "category_source": "category_source",
                    "is_unclassified": "is_unclassified",
                    "error_message": "error_message"
                }
                dict_key = key_map.get(pname, pname)
                cell.value = row_data.get(dict_key, "")
                
            # 6. Apply copied styling (borders, fonts) to the cell (preserve template fill)
            cached_style = physical_styles_cache.get(pname)
            if cached_style:
                cell.font = Font(
                    name=cached_style["font"].name,
                    size=cached_style["font"].size,
                    bold=cached_style["font"].bold,
                    italic=cached_style["font"].italic,
                    color=cached_style["font"].color,
                    underline=cached_style["font"].underline
                )
                cell.border = Border(
                    left=cached_style["border"].left,
                    right=cached_style["border"].right,
                    top=cached_style["border"].top,
                    bottom=cached_style["border"].bottom
                )
                
                # Check wrap alignment
                wrap_alignment = cached_style["alignment"].wrap_text or (pname in WRAP_COLS)
                cell.alignment = Alignment(
                    horizontal=cached_style["alignment"].horizontal,
                    vertical=cached_style["alignment"].vertical,
                    wrap_text=wrap_alignment
                )
                
                # Keep number format unless overridden
                if cell.number_format == 'General':
                    cell.number_format = cached_style["number_format"]
                    
    # 7. Apply Autofilter
    col_letter_last = get_column_letter(1 + len(active_columns))
    ws.auto_filter.ref = f"B3:{col_letter_last}{last_row_index}"
    
    # 8. Apply Freeze Panes
    ws.freeze_panes = "A4"
    
    # 9. Update all pivot cache range definitions
    for sheet_name in wb.sheetnames:
        ws_pivot = wb[sheet_name]
        for pivot in ws_pivot._pivots:
            pivot.cache.cacheSource.worksheetSource.ref = f"B3:{col_letter_last}{last_row_index}"
            
    # 10. Populate summary sheets (集計詳細 / 集計概要)
    populate_summary_sheets(wb, rows)
            
    # 11. Save workbook
    wb.save(output_path)

