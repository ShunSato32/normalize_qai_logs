import os
import json
import openpyxl
from openpyxl.styles import Alignment, Border, Side, PatternFill, Font
from openpyxl.utils import get_column_letter
from datetime import datetime
from typing import List, Dict, Any

CIRCLED_NUMS = ["①", "②", "➂", "④", "⑤", "⑥", "⑦", "⑧", "➈", "➉"]

def get_circled_num_doc_content(rank: int) -> str:
    if rank == 3:
        return "➂"  # U+2782
    elif rank == 9:
        return "➈"  # U+2788
    elif rank == 10:
        return "➉"  # U+2789
    else:
        return CIRCLED_NUMS[rank - 1]

def get_circled_num_score(rank: int) -> str:
    if rank == 3:
        return "③"  # U+2462
    elif rank == 9:
        return "➈"  # U+2788
    elif rank == 10:
        return "➉"  # U+2789
    else:
        return CIRCLED_NUMS[rank - 1]

# Build the complete headers list (110 columns total)
HEADERS = [
    # 15.1
    "output_row_id",
    "No.",
    "interaction_id",
    "message_id",
    "conversation_id",
    "reset_session_id",
    "reset_session_no",
    "interaction_no_in_session",
    "is_first_interaction_row",
    # 15.2
    "user_name",
    "team_name",
    "質問内容",
    "user_content_raw",
    "is_natural_language_query",
    "is_reset_request",
    "is_category_selection",
    "is_system_command",
    # 15.3
    "started_at_utc",
    "completed_at_utc",
    "started_at_jst",
    "completed_at_jst",
    "latency_sec",
    # 15.4
    "selected_function",
    "predicted_category",
    "user_selected_category",
    "fist_category",
    "final_category",
    "category_source",
    "is_unclassified",
    # 15.5
    "回答内容",
    "has_answer",
    "is_no_answer",
    "assistant_event_count",
    "has_error",
    "error_count",
    "error_message",
    # 15.6
    "retrieval_count",
    "retrieval_stored_count",
    "retrieval_truncated",
]

# Retrieval fixed columns (Columns 41 to 70)
for k in range(1, 11):
    doc_c = get_circled_num_doc_content(k)
    score_c = get_circled_num_score(k)
    HEADERS.append(f"参照文書{doc_c}")
    HEADERS.append(f"参考箇所{doc_c}")
    HEADERS.append(f"score{score_c}")

HEADERS.extend([
    # 15.8
    "feedback_id",
    "feedback_seq",
    "feedback_count",
    "feedback_rating",
    "feedback_comment",
    "feedback_at_utc",
    "feedback_at_jst",
    "is_latest_feedback",
    # 15.9
    "session_ends_with_reset",
    "source_files",
    "source_event_row_count",
    "source_event_types",
    "normalization_version",
    "normalization_warnings",
    # Existing Business / Evaluation columns
    "質問/回答\n分類",
    "対象/\n対象外",
])

for k in range(1, 11):
    doc_c = get_circled_num_doc_content(k)
    HEADERS.append(f"参照箇所{doc_c}")

HEADERS.extend([
    "教師データ有無",
    "キーワード1",
    "キーワード2",
    "キーワード3",
    "Hit判定",
    "回答判定",
    "正答判定",
    "分析コメント",
    "原因",
    "対応方針",
    "対応進捗",
    "Vantiq相談",
    "日付",
    "所属",
    "messageID"
])

WRAP_COLS = {
    "質問内容", "user_content_raw", "回答内容", "error_message", 
    "feedback_comment", "normalization_warnings", "分析コメント", "評価コメント"
}
for k in range(1, 11):
    doc_c = get_circled_num_doc_content(k)
    WRAP_COLS.add(f"参考箇所{doc_c}")

def parse_iso_datetime(dt_str: str) -> datetime:
    if not dt_str:
        return None
    try:
        # standard ISO string parsing, strip offset (e.g. +09:00, +00:00) to keep naive datetime local to that tz
        if '+' in dt_str:
            dt_str = dt_str.split('+')[0]
        if dt_str.endswith('Z'):
            dt_str = dt_str[:-1]
        return datetime.fromisoformat(dt_str)
    except Exception:
        return None

def write_integrated_to_excel(template_path: str, output_path: str, rows: List[Dict[str, Any]]):
    # Load template workbook
    wb = openpyxl.load_workbook(template_path, data_only=False)
    
    if "実施記録シート" not in wb.sheetnames:
        raise ValueError("Template Excel file must contain a worksheet named '実施記録シート'")
        
    ws = wb["実施記録シート"]
    
    # 1. Cache the styles from row 4 before clearing values.
    # We map from physical column name to its style, based on original HEADERS mapping.
    physical_styles_cache = {}
    for idx, orig_header in enumerate(HEADERS):
        col_idx = idx + 2
        if col_idx <= ws.max_column:
            cell = ws.cell(row=4, column=col_idx)
            physical_styles_cache[orig_header] = {
                "font": cell.font,
                "fill": cell.fill,
                "border": cell.border,
                "alignment": cell.alignment,
                "number_format": cell.number_format
            }

    # 2. Load column configuration dynamically
    config_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), "column_config.json")
    active_columns = []
    if os.path.exists(config_path):
        try:
            with open(config_path, "r", encoding="utf-8") as f:
                config_data = json.load(f)
            for item in config_data:
                pname = item.get("physical_name")
                jname = item.get("japanese_name") or pname
                if pname:
                    active_columns.append((pname, jname))
        except Exception as e:
            import sys
            print(f"Warning: Failed to load column_config.json: {e}. Using default headers.", file=sys.stderr)
            
    if not active_columns:
        active_columns = [(h, h) for h in HEADERS]
        
    # Map physical name to its active column letter in the JST sheet
    col_letters = {}
    for idx, (pname, jname) in enumerate(active_columns):
        col_idx = idx + 2
        col_letters[pname] = get_column_letter(col_idx)
        
    # 3. Clear all values from row 4 to the end of the sheet
    max_row = ws.max_row
    max_col = ws.max_column
    for r in range(4, max_row + 1):
        for c in range(1, max_col + 1):
            ws.cell(row=r, column=c).value = None
            
    # 4. Overwrite Row 3 headers from Column B (index 2) onwards
    for idx, (pname, jname) in enumerate(active_columns):
        col_idx = idx + 2
        cell = ws.cell(row=3, column=col_idx)
        cell.value = jname
        
    # 5. Populate rows from row 4 onwards
    last_row_index = 3 + len(rows)
    if len(rows) == 0:
        last_row_index = 4
        
    for row_num, row_data in enumerate(rows, start=4):
        for idx, (pname, jname) in enumerate(active_columns):
            col_idx = idx + 2
            cell = ws.cell(row=row_num, column=col_idx)
            
            # Write value or formula
            if pname == "No.":
                cell.value = f"=ROW(A{row_num})-3"
            elif pname == "対象/\n対象外":
                # References "質問/回答\n分類" (CC in default)
                q_a_class_col = col_letters.get("質問/回答\n分類", "CC")
                cell.value = f'=IF(OR({q_a_class_col}{row_num}="①", {q_a_class_col}{row_num}="⑤"), "〇", "×")'
            elif pname.startswith("参照箇所") and len(pname) > 4:
                circle = pname[4:]
                try:
                    rank = CIRCLED_NUMS.index(circle) + 1
                except ValueError:
                    if circle == "➂":
                        rank = 3
                    elif circle == "➈":
                        rank = 9
                    elif circle == "➉":
                        rank = 10
                    else:
                        rank = 1
                ref_desc_col = col_letters.get(f"参考箇所{circle}", "AP")
                kw1_col = col_letters.get("キーワード1", "CT")
                kw2_col = col_letters.get("キーワード2", "CU")
                kw3_col = col_letters.get("キーワード3", "CV")
                cell.value = f'=IF(OR(AND(${kw1_col}{row_num}<>"",ISNUMBER(SEARCH(${kw1_col}{row_num}, {ref_desc_col}{row_num}))),AND(${kw2_col}{row_num}<>"",ISNUMBER(SEARCH(${kw2_col}{row_num}, {ref_desc_col}{row_num}))),AND(${kw3_col}{row_num}<>"",ISNUMBER(SEARCH(${kw3_col}{row_num}, {ref_desc_col}{row_num})))),"〇","-")'
            elif pname == "Hit判定":
                check_cols = []
                for k in range(1, 11):
                    circle = get_circled_num_doc_content(k)
                    col_let = col_letters.get(f"参照箇所{circle}")
                    if col_let:
                        check_cols.append(f'{col_let}{row_num}="〇"')
                if check_cols:
                    cell.value = f'=IF(OR({",".join(check_cols)}),"〇","")'
                else:
                    cell.value = ""
            elif pname == "日付":
                start_jst_col = col_letters.get("started_at_jst", "U")
                cell.value = f'=TEXT({start_jst_col}{row_num},"yyyy/mm/dd")'
            elif pname == "所属":
                user_name_col = col_letters.get("user_name", "K")
                cell.value = f'=VLOOKUP({user_name_col}{row_num},社員マスタ!C:D,2,0)'
            elif pname == "messageID":
                cell.value = str(row_data.get("message_id", ""))
                cell.data_type = "s"
            elif pname in ("output_row_id", "interaction_id", "message_id", "conversation_id", "reset_session_id", "feedback_id"):
                cell.value = str(row_data.get(pname, ""))
                cell.data_type = "s"
            elif pname in ("started_at_utc", "completed_at_utc", "started_at_jst", "completed_at_jst", "feedback_at_utc", "feedback_at_jst"):
                val_str = row_data.get(pname, "")
                parsed_dt = parse_iso_datetime(val_str)
                cell.value = parsed_dt
                cell.number_format = 'yyyy/mm/dd hh:mm:ss'
            elif pname == "latency_sec":
                val = row_data.get(pname, "")
                if val != "" and val is not None:
                    cell.value = float(val)
                else:
                    cell.value = None
            elif "score" in pname:
                try:
                    circle = pname[5:]
                except Exception:
                    circle = ""
                rank = 1
                for idx_c, c in enumerate(CIRCLED_NUMS):
                    if c == circle:
                        rank = idx_c + 1
                        break
                if circle == "➂":
                    rank = 3
                elif circle == "➈":
                    rank = 9
                elif circle == "➉":
                    rank = 10
                col_name = f"retrieval_{rank:02d}"
                json_str = row_data.get(col_name)
                if json_str:
                    try:
                        data = json.loads(json_str)
                        score = data.get("score")
                        if score is not None and score != "":
                            cell.value = float(score)
                            cell.number_format = '0.000000'
                        else:
                            cell.value = None
                    except Exception:
                        cell.value = None
                else:
                    cell.value = None
            elif pname.startswith("参照文書") and len(pname) > 4:
                circle = pname[4:]
                rank = 1
                for idx_c, c in enumerate(CIRCLED_NUMS):
                    if c == circle:
                        rank = idx_c + 1
                        break
                if circle == "➂":
                    rank = 3
                elif circle == "➈":
                    rank = 9
                elif circle == "➉":
                    rank = 10
                col_name = f"retrieval_{rank:02d}"
                json_str = row_data.get(col_name)
                if json_str:
                    try:
                        data = json.loads(json_str)
                        filename = data.get("filename") or data.get("fileName") or data.get("display_name") or ""
                        cell.value = filename
                    except Exception:
                        cell.value = ""
                else:
                    cell.value = ""
            elif pname.startswith("参考箇所") and len(pname) > 4:
                circle = pname[4:]
                rank = 1
                for idx_c, c in enumerate(CIRCLED_NUMS):
                    if c == circle:
                        rank = idx_c + 1
                        break
                if circle == "➂":
                    rank = 3
                elif circle == "➈":
                    rank = 9
                elif circle == "➉":
                    rank = 10
                col_name = f"retrieval_{rank:02d}"
                json_str = row_data.get(col_name)
                if json_str:
                    try:
                        data = json.loads(json_str)
                        cell.value = data.get("content", "")
                    except Exception:
                        cell.value = ""
                else:
                    cell.value = ""
            elif pname in (
                "feedback_seq", "feedback_count", "retrieval_count", "retrieval_stored_count", 
                "retrieval_truncated", "is_first_interaction_row", "is_latest_feedback", 
                "is_natural_language_query", "is_reset_request", "is_category_selection", 
                "is_system_command", "session_ends_with_reset", "has_answer", "is_no_answer", 
                "assistant_event_count", "has_error", "error_count", "source_event_row_count"
            ):
                val = row_data.get(pname)
                if val != "" and val is not None:
                    cell.value = int(val)
                else:
                    cell.value = 0
            else:
                key_map = {
                    "predicted_category": "predicted_category",
                    "user_selected_category": "user_selected_category",
                    "回答内容": "回答内容",
                    "fist_category": "fist_category",
                    "final_category": "final_category",
                    "category_source": "category_source",
                    "is_unclassified": "is_unclassified",
                    "error_message": "error_message"
                }
                dict_key = key_map.get(pname, pname)
                cell.value = row_data.get(dict_key, "")
                
            # 6. Apply copied styling (borders, fonts, fills) to the cell
            cached_style = physical_styles_cache.get(pname)
            if cached_style:
                cell.font = Font(
                    name=cached_style["font"].name,
                    size=cached_style["font"].size,
                    bold=cached_style["font"].bold,
                    italic=cached_style["font"].italic,
                    color=cached_style["font"].color,
                    underline=cached_style["font"].underline
                )
                cell.fill = PatternFill(
                    fill_type=cached_style["fill"].fill_type,
                    start_color=cached_style["fill"].start_color,
                    end_color=cached_style["fill"].end_color
                )
                cell.border = Border(
                    left=cached_style["border"].left,
                    right=cached_style["border"].right,
                    top=cached_style["border"].top,
                    bottom=cached_style["border"].bottom
                )
                
                # Check wrap alignment
                wrap_alignment = cached_style["alignment"].wrap_text or (pname in WRAP_COLS)
                cell.alignment = Alignment(
                    horizontal=cached_style["alignment"].horizontal,
                    vertical=cached_style["alignment"].vertical,
                    wrap_text=wrap_alignment
                )
                
                # Keep number format unless overridden
                if cell.number_format == 'General':
                    cell.number_format = cached_style["number_format"]
                    
    # 7. Apply Autofilter
    col_letter_last = get_column_letter(1 + len(active_columns))
    ws.auto_filter.ref = f"B3:{col_letter_last}{last_row_index}"
    
    # 8. Apply Freeze Panes
    ws.freeze_panes = "A4"
    
    # 9. Update all pivot cache range definitions
    for sheet_name in wb.sheetnames:
        ws_pivot = wb[sheet_name]
        for pivot in ws_pivot._pivots:
            pivot.cache.cacheSource.worksheetSource.ref = f"B3:{col_letter_last}{last_row_index}"
            
    # 10. Save workbook
    wb.save(output_path)
