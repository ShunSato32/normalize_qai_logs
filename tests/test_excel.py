import os
import openpyxl
import pytest
from excel_writer import write_integrated_to_excel, HEADERS

def test_excel_export_and_formatting(tmp_path):
    # Setup paths
    script_dir = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
    template_path = os.path.join(script_dir, "templates", "【生成AI】実施記録分析シート_20260610132229.xlsx")
    
    if not os.path.exists(template_path):
        pytest.skip("Template Excel file not found for testing.")
        
    output_path = os.path.join(tmp_path, "integrated_output.xlsx")
    
    # Create mock integrated rows
    mock_rows = [
        {
            "unique_row_id": "CONV001-MSG001-FB001",
            "interaction_id": "CONV001-MSG001",
            "message_id": "MSG001",
            "conversation_id": "CONV001",
            "reset_session_id": "CONV001-S001",
            "reset_session_no": 1,
            "interaction_no_in_session": 1,
            "is_first_interaction_row": 1,
            "user_name": "User A",
            "team_name": "Team A",
            "question": "Hello QA",
            "user_content_raw": "Hello QA",
            "is_natural_language_query": 1,
            "is_reset_request": 0,
            "is_category_selection": 0,
            "is_system_command": 0,
            "started_at_jst": "2026-06-02T21:00:00+09:00",
            "completed_at_jst": "2026-06-02T21:00:02+09:00",
            "latency_sec": "2.000",
            "selected_function": "func1",
            "predicted_category": "制度・運用 > 服務",
            "user_selected_category": "",
            "fist_category": "制度・運用",
            "final_category": "制度・運用 > 服務",
            "category_source": "predicted",
            "is_unclassified": 0,
            "answer": "This is a reply",
            "has_answer": 1,
            "is_no_answer": 0,
            "assistant_event_count": 1,
            "has_error": 0,
            "error_count": 0,
            "error_message": "",
            "retrieval_count": 1,
            "retrieval_stored_count": 1,
            "retrieval_truncated": 0,
            "retrieval_01": '{"rank": 1, "score": 0.954321, "filename": "doc1.pdf", "content": "excerpt"}',
            "feedback_id": "CONV001-MSG001-FB001",
            "feedback_seq": 1,
            "feedback_count": 1,
            "feedback_rating": "good",
            "feedback_comment": "helpful",
            "feedback_at_jst": "2026-06-02T21:05:00+09:00",
            "is_latest_feedback": 1,
            "session_ends_with_reset": 0,
            "source_files": "file1.csv",
            "source_event_row_count": 4,
            "source_event_types": "user > assistant > tool > feedback",
            "normalization_version": "1.0.0",
            "normalization_warnings": ""
        }
    ]
    
    # Run exporter
    write_integrated_to_excel(template_path, output_path, mock_rows)
    
    # Load back the output to verify
    wb_out = openpyxl.load_workbook(output_path, data_only=False)
    
    # Test case 31: Verify worksheets in output workbook are preserved
    expected_sheets = [
        '実施記録シート', 
        '簡易集計_投入数(日付×カテゴリ)', 
        '簡易集計_投入数(日付×カテゴリ自動設定)', 
        '簡易集計_投入数(日付×ユーザ評価)', 
        '簡易集計_参加者数(日付×部門・担当)', 
        '社員マスタ'
    ]
    for sheet in expected_sheets:
        assert sheet in wb_out.sheetnames, f"Worksheet {sheet} was not preserved!"
        
    ws_out = wb_out["実施記録シート"]
    
    # Load config to find actual columns position dynamically
    import json
    config_file_path = os.path.join(script_dir, "config", "column_config.json")
    with open(config_file_path, "r", encoding="utf-8") as f:
        config_data = json.load(f)
    active_pnames = [item["physical_name"] for item in config_data]

    # Test case 32: Unique header names in Row 3
    row3_values = [cell.value for cell in ws_out[3] if cell.value is not None]
    assert len(row3_values) == len(active_pnames), "Headers length mismatch!"
    assert len(row3_values) == len(set(row3_values)), "Duplicate headers found in row 3!"
    
    # Test case 33: ID columns written as string type

    id_pnames = ["unique_row_id", "interaction_id", "message_id", "conversation_id", "reset_session_id"]
    id_cols = [active_pnames.index(name) + 2 for name in id_pnames if name in active_pnames]
    for col in id_cols:
        cell = ws_out.cell(row=4, column=col)
        assert cell.data_type == 's', f"ID column {col} was not saved as a string!"
        
    # Test case 34: DateTime columns written as datetime object
    started_at_jst_col = active_pnames.index("started_at_jst") + 2
    cell_dt = ws_out.cell(row=4, column=started_at_jst_col)
    import datetime
    assert isinstance(cell_dt.value, datetime.datetime), "DateTime column was not parsed into datetime cell!"
    assert cell_dt.number_format == 'yyyy/mm/dd hh:mm:ss', "DateTime cell has incorrect format mask!"
    
    # Test case 35: Score column written as numeric float
    score_1_col = active_pnames.index("score_1") + 2
    cell_score = ws_out.cell(row=4, column=score_1_col)
    assert isinstance(cell_score.value, float), "Score was not saved as numeric float!"
    assert abs(cell_score.value - 0.954321) < 1e-6
    assert cell_score.number_format == '0.000000', "Score format mask was not set!"
    
    # Test case 36: Autofilter and Freeze Panes are active
    assert ws_out.auto_filter.ref is not None, "Autofilter was not set!"
    assert ws_out.auto_filter.ref.startswith("B3:"), "Autofilter range start mismatch!"
    assert ws_out.freeze_panes == "A4", "Freeze panes was not set to A4!"
    
    # Verify pivot table range is updated
    ws_pivot = wb_out['簡易集計_投入数(日付×カテゴリ)']
    p = ws_pivot._pivots[0]
    from openpyxl.utils import get_column_letter
    col_letter_last = get_column_letter(1 + len(active_pnames))
    assert p.cache.cacheSource.worksheetSource.ref == f"B3:{col_letter_last}4", "Pivot Table cache ref range was not updated!"


def test_custom_column_config(tmp_path):
    import json
    # Setup paths
    script_dir = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
    template_path = os.path.join(script_dir, "templates", "【生成AI】実施記録分析シート_20260610132229.xlsx")
    
    if not os.path.exists(template_path):
        pytest.skip("Template Excel file not found for testing.")
        
    output_path = os.path.join(tmp_path, "custom_integrated_output.xlsx")
    
    # Create custom config data
    custom_config = [
        {"physical_name": "user_name", "japanese_name": "カスタムユーザー名"},
        {"physical_name": "team_name", "japanese_name": "カスタムチーム名"},
        {"physical_name": "No.", "japanese_name": "カスタムNo"}
    ]
    
    config_file_path = os.path.join(script_dir, "config", "column_config.json")
    original_exists = os.path.exists(config_file_path)
    
    # Read original if it exists
    original_content = None
    if original_exists:
        with open(config_file_path, "r", encoding="utf-8") as f:
            original_content = f.read()
            
    # Write custom config
    with open(config_file_path, "w", encoding="utf-8") as f:
        json.dump(custom_config, f)
        
    try:
        mock_rows = [
            {
                "team_name": "テストチーム",
                "user_name": "テストユーザー"
            }
        ]
        # Run exporter
        write_integrated_to_excel(template_path, output_path, mock_rows)
        
        # Verify columns in the generated workbook
        wb_out = openpyxl.load_workbook(output_path, data_only=False)
        ws_out = wb_out["実施記録シート"]
        
        # Row 3 should have our custom headers in block order
        headers_row3 = [ws_out.cell(row=3, column=c).value for c in range(2, 5)]
        assert headers_row3 == ["カスタムユーザー名", "カスタムチーム名", "カスタムNo"]
        
        # Row 4 values should be mapped correctly to the sorted columns
        # Col 2 (B) should be "テストユーザー"
        # Col 3 (C) should be "テストチーム"
        # Col 4 (D) should be formula "=ROW(A4)-3"
        assert ws_out.cell(row=4, column=2).value == "テストユーザー"
        assert ws_out.cell(row=4, column=3).value == "テストチーム"
        assert ws_out.cell(row=4, column=4).value == "=ROW(A4)-3"
        
    finally:
        # Restore original config
        if original_exists:
            with open(config_file_path, "w", encoding="utf-8") as f:
                f.write(original_content)
        else:
            if os.path.exists(config_file_path):
                os.remove(config_file_path)


def test_system_command_formatting(tmp_path):
    import json
    script_dir = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
    template_path = os.path.join(script_dir, "templates", "【生成AI】実施記録分析シート_20260610132229.xlsx")
    if not os.path.exists(template_path):
        pytest.skip("Template Excel file not found for testing.")
        
    output_path = os.path.join(tmp_path, "syscmd_output.xlsx")
    mock_rows = [
        {
            "unique_row_id": "SYS001",
            "is_system_command": 1,
            "question": "こんにちは"
        }
    ]
    write_integrated_to_excel(template_path, output_path, mock_rows)
    
    wb_out = openpyxl.load_workbook(output_path, data_only=False)
    ws_out = wb_out["実施記録シート"]
    
    # Read active columns from config
    config_file_path = os.path.join(script_dir, "config", "column_config.json")
    with open(config_file_path, "r", encoding="utf-8") as f:
        config_data = json.load(f)
    active_pnames = [item["physical_name"] for item in config_data]
    
    qa_col = active_pnames.index("qa_classification") + 2 if "qa_classification" in active_pnames else None
    target_col = active_pnames.index("is_target") + 2 if "is_target" in active_pnames else None
    
    if qa_col and target_col:
        assert ws_out.cell(row=4, column=qa_col).value == "④集計対象外"
        assert ws_out.cell(row=4, column=target_col).value == "×"

